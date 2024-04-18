import pandas as pd
import json
import psycopg2
from openpyxl import load_workbook
import os
import inspect


def convert_data_type(value):
    """
    Attempts to convert a string value to number, date, or currency.
    Returns the original value if conversion fails.
    """
    if value is None:
        return None  # You can return a different default value if needed
    
    try:
        # Try float conversion
        return float(value.replace(",", ""))
    except ValueError:
        pass
    try:
        # Try date conversion (adjust format as needed for your data)
        return pd.to_datetime(value, format="%Y-%m-%d")
    except ValueError:
        pass
    # No conversion successful, return original value
    return value

def unmerge_cells(sheet):
    """
    Unmerges merged cells in a worksheet. 
    - For same-row merges, copies data to unmerged columns.
    - For multi-row merges (unsupported), logs a warning.
    """
    for merge_range in sheet.merged_cells.ranges:
        # Check if coord contains four values (expected for merged cells)
        if len(merge_range.coord.split(':')) != 4:
            continue  # Skip non-merged cells

        try:
            # Unpack coordinates if format is valid
            start_row, start_col, end_row, end_col = merge_range.coord.split(':')
            start_row, start_col, end_row, end_col = int(start_row), int(start_col[0]), int(end_row), int(end_col[0])
        except ValueError:
            print(f"Error parsing coordinates for merged cell {merge_range.coord}. Skipping.")
            continue

        # Check if merging spans multiple rows (unsupported)
        if start_row != end_row:
            print(f"Warning: Multi-row merged cells detected at {merge_range.coord}. Not supported yet.")
            continue

        # Get value from top-left cell (assuming data is present there)
        reference_value = sheet.cell(row=start_row, column=start_col).value

        for col in range(start_col, end_col + 1):
            # Fill empty cells within the same row with reference value
            if sheet.cell(row=start_row, column=col).value is None:
                sheet.cell(row=start_row, column=col).value = reference_value

def process_excel_file(filename, conn, table_headers, use_excel_headers=True):
    """
    Reads an Excel sheet, converts data, creates JSON, and persists data to a Postgres table.

    Args:
        filename (str): Name of the Excel file.
        conn (psycopg2.connect): Connection object to the Postgres database.
        table_headers (list): Static list of table headers (if use_excel_headers is False).
        use_excel_headers (bool, optional): Flag to control header processing. Defaults to True.
    """
    try:
        cursor = conn.cursor()
        file_path = "C:\\CCC\\postpaid\\"

        # Load workbook and unmerge cells
        wb = load_workbook(filename=os.path.join(file_path, filename), data_only=True)
        sheet = wb.active
        unmerge_cells(sheet)  # Assume unmerge_cells is your custom function

        data_list = list(sheet.values)
        print("LIST DATA:: ")
        print(data_list)
        # Transpose data directly from list (handling missing values)
        transposed_data = []
        for col in range(len(data_list[0])):  # Use length of first row for columns
            transposed_data.append([row[col] if len(row) > col else None for row in data_list[0:]])

        # Ensure all rows in transposed data have the same length
        max_length = max(len(row) for row in transposed_data)
        for row in transposed_data:
            row.extend([None] * (max_length - len(row)))  # Pad with None if shorter
        filtered_transposed_data = [row for row in transposed_data if any(val is not None for val in row)]
        # Determine table headers based on flag (after transpose)
        table_headers = None
        if use_excel_headers:
            # Use first row from transposed data (excluding integer sequence header)
            table_headers = filtered_transposed_data[0][0:]
        else:
            # Handle static headers (implementation depends on your logic)
            pass  # Replace with your static header logic

        # Create pandas DataFrame with transposed data and headers
        #print("TABLE HEADERS :: ")
        #print(table_headers)
        #print("TRANSPOSED DATA ::")
        #print(transposed_data)
        df = pd.DataFrame(filtered_transposed_data[1:], columns=table_headers)
        print("VALUE LIST::")
        print(df)

        # Backup existing data (implementation depends on your database)
        # ... (replace with your backup logic)

        # Convert DataFrame to JSON and persist to database
        replacement_char = "_"  # You can change this to "-" or use camel case logic
        # Generate table name without spaces (replace with your logic)
        table_name = os.path.splitext(filename)[0]  # Get filename without extension
        table_name = table_name.replace(" ", replacement_char)  # Replace spaces

            # Check if table exists
        cursor.execute(f"""
            SELECT EXISTS (
                SELECT * FROM information_schema.tables
                WHERE table_schema = current_schema()
                AND table_name = '{table_name}'
            );
        """)
        table_exists = cursor.fetchone()[0]
        cursor.close()

        # Truncate table only if it exists
        if table_exists:
            cursor = conn.cursor()
            cursor.execute(f"TRUNCATE TABLE {table_name};")
            conn.commit()
            cursor.close()
        else:
            print(f"Table '{table_name}' does not exist. Skipping truncation.")
        
        # Replace spaces with underscores in column names
        df.columns = df.columns.str.replace(' ', '_')

        # Get column names from DataFrame
        column_names = df.columns.tolist()  # Skip first column (integer sequence, if present)

        print(column_names)

        # Dynamically generate placeholders using string formatting
        placeholders = ', '.join(['%s'] * len(column_names))
        
        # Insert data into database using psycopg2 (no JSON conversion needed)
        # Insert data into database table (replace "your_table_name" with actual table name)
        cursor = conn.cursor()
        #insert_stmt = cursor.mogrify(f"INSERT INTO {table_name} ({','.join(table_headers)}) VALUES ({placeholders})")
        #print(insert_stmt)
        #cursor.executemany(insert_stmt, df)

        # Construct insert statement
        insert_stmt = f"""INSERT INTO {table_name} ({','.join(column_names)}) VALUES ({placeholders})"""

        # Prepare data for insertion
        values = df.values.tolist()  # List of lists with row values

        # Execute insert statement using executemany
        cursor.executemany(insert_stmt, values)

        conn.commit()
    except IndexError as e:
        print("Error: String index out of range!")
        print(f"Failing line: {inspect.currentframe().f_lineno}")
        print(f"Potentially problematic variable: table_headers = {table_headers}")
    except (Exception, psycopg2.Error) as error:
        print("Error while processing file or inserting data:", error)
    finally:
        if conn is not None:
            cursor.close()
            conn.close()

def main():
    # Replace with your connection details
    conn = psycopg2.connect(dbname="postgres", user="postgres", password="admin", host="localhost", port="5432")

    # Configure static table headers (replace with your actual headers)
    static_table_headers = ["col1", "col2", "col3"]

    # Specify the file path containing Excel files
    file_path = "C:\\CCC\\postpaid\\"

    # Get list of Excel files in the directory
    excel_files = [f for f in os.listdir(file_path) if f.endswith(".xlsx")]

    for filename in excel_files:
        # Process each Excel file using process_excel_file function
        process_excel_file(filename, conn, static_table_headers, use_excel_headers=True)  # Change use_excel_headers as needed

    conn.close()
    print("Data processing completed!")

if __name__ == "__main__":
    main()
